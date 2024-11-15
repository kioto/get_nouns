"""文書ファイル（テキスト、Excel）から名詞一覧を表示
"""
import sys
import re
from pathlib import Path
from janome.tokenizer import Tokenizer
import openpyxl


USER_DICT_CSV = './user_dict.csv'  # ユーザー辞書
DST_DICT_CSV = '/tmp/_dict.csv'    # 中間辞書ファイル
POS1_TYPE = ['一般', '固有名詞', '*']


def get_part_of_speech(token, pos_type=0):
    """トークンから品詞情報を取得
    pos_type=0: 品詞
    pos_type=1: 品詞細分類1
    pos_type=2: 品詞細分類2
    """
    return token.part_of_speech.split(',')[pos_type]


def is_noun(token):
    """トークンが名詞（あるいはカスタム名詞）かを判定
    """
    pos0 = get_part_of_speech(token)
    pos1 = get_part_of_speech(token, 1)
    return pos0 in ['名詞', u'カスタム名詞'] and pos1 in POS1_TYPE


def is_text_file(filepath):
    """ファイルがテキストファイルか判定
    """
    return filepath.suffix.lower() == '.txt'


def is_excel_file(filepath):
    """ファイルがExcelファイルか判定
    """
    return filepath.suffix.lower() == '.xlsx'


def entry_nouns(noun_map, tokens):
    """トークンリストから名詞トークンを名詞mapに登録する
    """
    for token in tokens:
        if is_noun(token) and token.surface not in noun_map:
            noun_map[token.surface] = token


def get_nouns_from_text_file(filename, tokenizer):
    """テキストファイルから名詞一覧を返す
    """
    noun_map = {}
    with open(filename) as f:
        for line in f.readlines():
            entry_nouns(noun_map, tokenizer.tokenize(line))

    return noun_map


def get_cell_texts(sheet):
    """シートからセルのテキストリストを取り出す
    """
    texts = []
    for row in [sheet[i] for i in range(1, sheet.max_row+1)]:
        # 行単位で取り出し
        for cell in row:
            # セル単位で取り出し
            if cell.value is not None and isinstance(cell.value, str):
                texts.append(cell.value)
    return texts


def get_nouns_from_excel_file(filename, tokenizer):
    """Excelファイルから名詞一覧を返す
    """
    noun_map = {}
    wb = openpyxl.load_workbook(filename)
    for sheet_name in wb.sheetnames:
        # シート名も対象
        entry_nouns(noun_map, tokenizer.tokenize(sheet_name))

        # シートの内容
        for cell_text in get_cell_texts(wb[sheet_name]):
            entry_nouns(noun_map, tokenizer.tokenize(cell_text))

    return noun_map


def main(filename, dict_file):
    in_file = Path(filename)
    if not in_file.exists():
        print(f'{in_file}: No such file or directory')
        return

    # トーカナイザー生成
    tokenizer = None
    if Path(dict_file).exists():
        tokenizer = Tokenizer(dict_file,
                              udic_type='simpledic', udic_enc='utf8')
    else:
        tokenizer = Tokenizer()

    # 名詞一覧を取得
    noun_map = {}
    if is_text_file(in_file):
        noun_map = get_nouns_from_text_file(in_file, tokenizer)
    elif is_excel_file(in_file):
        noun_map = get_nouns_from_excel_file(in_file, tokenizer)

    # 名詞一覧を表示
    for token in noun_map.values():
        # XXX	名詞,固有名詞,地域,一般,*,*,XXX,YYY,YYY
        # とある場合、
        # XXX	名詞,固有名詞,地域,一般
        # のみを表示する
        msg = re.sub(r'^(.+\t)([^,]+),([^,]+),([^,]+),([^,]+),(.*$)',
                     r'\1\2,\3,\4,\5', str(token))

        print(msg)


def make_dict_file(dict_files, dst_dict_file):
    """辞書ファイルの作成
    dict_filesのファイルから注釈行を削除し、dst_dict_fileに書き込む。
    """
    with open(dst_dict_file, 'w') as fout:
        for dict_file in dict_files:
            if not Path(dict_file).exists():
                print(f'ERROR: {dict_file}: No such file or directory')
                exit()

            with open(dict_file) as fin:
                for line in fin.readlines():
                    if re.match(r'^#', line):
                        # 注釈行をスキップ
                        continue
                    fout.write(line)


if __name__ == '__main__':
    if len(sys.argv) == 1:
        print(f'Usage: python {sys.argv[0]} '
              '[-d <dict-file>] <text-file> | <excel-file>')
        exit()

    # option
    filename = ''
    is_dict_opt = False
    dict_files = []
    for arg in sys.argv[1:]:
        if is_dict_opt:
            is_dict_opt = False
            dict_files.append(arg)
        else:
            if arg == '-d':
                is_dict_opt = True
            else:
                filename = arg

    # 中間辞書ファイルが残っていれば削除
    if Path(DST_DICT_CSV).exists():
        Path(DST_DICT_CSV).unlink()

    if Path(USER_DICT_CSV).exists():
        # ユーザ辞書があれば追加
        dict_files.append(USER_DICT_CSV)

    # 辞書を中間辞書ファイルに集約
    if dict_files:
        make_dict_file(dict_files, DST_DICT_CSV)

    # メインルーチン
    main(filename, DST_DICT_CSV)

    # 中間辞書ファイルを削除
    if Path(DST_DICT_CSV).exists():
        Path(DST_DICT_CSV).unlink()
